# services/exports.py
from __future__ import annotations
import csv
from datetime import datetime
from typing import List, Dict, Any, Iterable

# openpyxl é o writer usado pelo Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# -----------------------------
# Utilidades
# -----------------------------
def _to_date(value: Any):
    """
    Converte 'dd/mm/aaaa' -> datetime.date.
    Retorna None se não conseguir converter.
    """
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for sep in ("/", "-", "."):
        if sep in s and len(s) >= 8:
            try:
                if sep == "/":
                    d, m, y = s.split("/")
                else:
                    y, m, d = s.split(sep)
                d, m, y = int(d), int(m), int(y)
                return datetime(y, m, d).date()
            except Exception:
                pass
    return None


def _is_number(value: Any) -> bool:
    try:
        float(str(value).replace(",", "."))
        return True
    except Exception:
        return False


# -----------------------------
# CSV
# -----------------------------
def export_csv(headers: List[str], rows: Iterable[Dict[str, Any]], path: str) -> None:
    """
    Exporta CSV simples em UTF-8 com cabeçalho.
    rows: iterável de dicts com as chaves == headers
    """
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in headers})


# -----------------------------
# XLSX Estilizado
# -----------------------------
def export_xlsx(headers: List[str], rows: Iterable[Dict[str, Any]], path: str) -> None:
    """
    Gera .xlsx com visual moderno:
      - Cabeçalho negrito, fundo cinza (#DDDDDD), centralizado, borda inferior.
      - Linhas com altura ~19, zebra #F9F9F9.
      - Fonte Calibri 11.
      - Texto centralizado vertical e horizontal (números/datas à direita).
      - Largura de coluna = conteúdo + 20%.
      - Datas 'dd/mm/aaaa'; valores 'R$ #.##0,00'.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Fonte e estilos básicos
    base_font = Font(name="Calibri", size=11)
    ws.sheet_view.showGridLines = False

    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_bottom = Border(bottom=thin)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    zebra_fill = PatternFill("solid", fgColor="F9F9F9")

    # Cabeçalho
    ws.append(headers)
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.alignment = align_center
        c.fill = header_fill
        c.border = border_bottom
    ws.row_dimensions[1].height = 22

    # Linhas
    max_width = [len(h) for h in headers]

    row_index = 2
    for r in rows:
        values = [r.get(h, "") for h in headers]
        ws.append(values)

        # zebra + altura
        ws.row_dimensions[row_index].height = 19
        if row_index % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_index, column=col_idx).fill = zebra_fill

        # formatação por coluna
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_index, column=col_idx)
            cell.font = base_font
            cell.border = border_all

            val = cell.value

            # Datas (se for string 'dd/mm/aaaa' converte)
            d = _to_date(val)
            if d:
                cell.value = d
                cell.number_format = "dd/mm/yyyy"
                cell.alignment = align_right
            # Moeda (se parece numérico e cabeçalho indica preço/valor)
            elif _is_number(val) and any(k in h.lower() for k in ("preço", "valor", "total")):
                try:
                    cell.value = float(str(val).replace(",", "."))
                except Exception:
                    pass
                cell.number_format = r'R$ #,##0.00'
                cell.alignment = align_right
            # Telefones/CNPJ/CPF permanecem texto centralizado
            elif any(k in h.lower() for k in ("cnpj", "cpf", "telefone", "celular", "agência", "conta")):
                cell.alignment = align_center
            # Números “puros”
            elif _is_number(val):
                cell.alignment = align_right
            else:
                cell.alignment = align_center

            # medir largura
            txt = str(val) if val is not None else ""
            max_width[col_idx - 1] = max(max_width[col_idx - 1], len(txt))

        row_index += 1

    # Largura de coluna (+20%)
    for i, w in enumerate(max_width, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = max(10, int(w * 1.2))

    # Borda externa leve (já temos bordas nas células)
    # Salva
    wb.save(path)


# -----------------------------
# Ajuda para gerar “modelos” (planilhas de exemplo)
# -----------------------------
def export_modelo_xlsx(title: str, headers: List[str], example_rows: List[Dict[str, Any]], path: str) -> None:
    """
    Gera um arquivo de exemplo bem formatado com 1–3 linhas de exemplo.
    """
    export_xlsx(headers, example_rows, path)
